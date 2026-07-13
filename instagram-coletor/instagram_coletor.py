"""
Instagram Coletor — baixa midia e metadados do Instagram usando gallery-dl.

Interface grafica (Tkinter). Tres modos:
  1. Links      — cola links de post / carrossel / reel e baixa a midia.
2. Perfil     — digita @perfil e baixa feed, reels, stories, destaques, marcados.
  3. Planilha   — gera .xlsx com legenda, data, curtidas, hashtags e link de cada post.

Autenticacao: o Instagram exige login para quase tudo. A ferramenta puxa os
cookies de um navegador ja logado (Chrome, Firefox, Edge, Brave, Opera) ou de
um arquivo cookies.txt.
"""

import logging
import os
import queue
import re
import sys
import threading
import traceback
import webbrowser
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# ---------------------------------------------------------------------------
# Dependencias externas
# ---------------------------------------------------------------------------
try:
    from gallery_dl import config as gdl_config
    from gallery_dl import job as gdl_job
    from gallery_dl import exception as gdl_exception
    from gallery_dl import cookies as gdl_cookies
except ImportError:  # pragma: no cover
    gdl_config = gdl_job = gdl_exception = gdl_cookies = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None


APP_NAME = "Instagram Coletor"
VERSION = "1.0.0"

NAVEGADORES = [
    "Chrome",
    "Firefox",
    "Edge",
    "Brave",
    "Opera",
    "Chromium",
    "Vivaldi",
    "Arquivo cookies.txt",
    "Nenhum (so posts publicos)",
]

SEM_LOGIN = "Nenhum (so posts publicos)"
ARQUIVO_COOKIES = "Arquivo cookies.txt"


def chave_navegador(rotulo: str) -> str:
    """Converte o rotulo da tela na chave que o gallery-dl entende."""
    return rotulo.split("(")[0].strip().lower()


SECOES_PERFIL = [
    ("posts", "Feed (posts e carrosseis)"),
    ("reels", "Reels"),
    ("stories", "Stories (24h)"),
    ("highlights", "Destaques"),
    ("tagged", "Marcados (fotos onde foi marcado)"),
    ("avatar", "Foto de perfil (alta resolucao)"),
]


# ---------------------------------------------------------------------------
# Ponte de log: o gallery-dl reporta erro pelo logging, nao por excecao.
# Sem isto, falha de login vira "0 itens" em silencio.
# ---------------------------------------------------------------------------
class PonteDeLog(logging.Handler):
    def __init__(self, escrever):
        super().__init__(level=logging.INFO)
        self.escrever = escrever

    def emit(self, record):
        try:
            msg = record.getMessage()
        except Exception:  # noqa: BLE001
            return
        nivel = record.levelname.upper()
        if nivel in ("ERROR", "CRITICAL"):
            self.escrever(f"  [ERRO] {msg}")
        elif nivel == "WARNING":
            self.escrever(f"  [aviso] {msg}")
        else:
            self.escrever(f"  {msg}")


def instalar_ponte_de_log(escrever):
    raiz = logging.getLogger()
    for h in list(raiz.handlers):
        if isinstance(h, PonteDeLog):
            raiz.removeHandler(h)
    ponte = PonteDeLog(escrever)
    raiz.addHandler(ponte)
    raiz.setLevel(logging.INFO)
    return ponte


# ---------------------------------------------------------------------------
# Diagnostico de login
# ---------------------------------------------------------------------------
def testar_login(navegador, caminho_cookies):
    """Diz, em portugues claro, se o cookie de sessao do Instagram foi achado."""
    if navegador == SEM_LOGIN:
        return False, ("Voce escolheu 'Nenhum'. O Instagram exige login para quase "
                       "tudo.\n\nEscolha o navegador em que voce esta logado no "
                       "Instagram.")

    if navegador == ARQUIVO_COOKIES:
        if not caminho_cookies or not Path(caminho_cookies).exists():
            return False, "Selecione o arquivo cookies.txt primeiro."
        texto = Path(caminho_cookies).read_text(errors="ignore")
        if "sessionid" in texto:
            return True, "Login OK. O arquivo tem o cookie de sessao do Instagram."
        return False, ("O arquivo nao tem o cookie 'sessionid' do Instagram.\n\n"
                       "Ele precisa ser exportado com a conta logada no Instagram.")

    nome = chave_navegador(navegador)
    try:
        jar = gdl_cookies.load_cookies((nome, None, None, None, ".instagram.com"))
    except Exception as e:  # noqa: BLE001
        erro = str(e)
        dica = ""
        if "permission" in erro.lower() or "operation not permitted" in erro.lower():
            dica = "\n\nO macOS bloqueou a leitura. Tente o Chrome ou o Firefox."
        return False, f"Nao consegui ler os cookies do {navegador}.\n\n{erro}{dica}"

    nomes = {c.name for c in jar}
    total = len(nomes)

    if "sessionid" in nomes:
        return True, (f"Login OK. Achei o cookie de sessao do Instagram no "
                      f"{navegador} ({total} cookies).")

    if total == 0:
        return False, (f"Nenhum cookie do Instagram encontrado no {navegador}.\n\n"
                       f"Abra o {navegador}, entre em instagram.com e faca login. "
                       f"Depois teste de novo.")

    return False, (f"Achei {total} cookies do Instagram no {navegador}, mas nao o "
                   f"'sessionid' (o que prova o login).\n\n"
                   f"Abra o {navegador}, entre em instagram.com e confirme que voce "
                   f"esta logado. Depois teste de novo.")


# ---------------------------------------------------------------------------
# Cancelamento cooperativo
# ---------------------------------------------------------------------------
CANCELAR = threading.Event()


class Cancelado(Exception):
    pass


def _fabricar_job_download(log_fn, progresso_fn):
    """Cria uma subclasse de DownloadJob que respeita o cancelamento."""

    class JobCancelavel(gdl_job.DownloadJob):
        def handle_url(self, url, kwdict):
            if CANCELAR.is_set():
                raise Cancelado()
            super().handle_url(url, kwdict)
            nome = kwdict.get("filename") or kwdict.get("post_shortcode") or "arquivo"
            ext = kwdict.get("extension", "")
            progresso_fn(f"{nome}.{ext}" if ext else str(nome))

        def handle_queue(self, url, kwdict):
            if CANCELAR.is_set():
                raise Cancelado()
            return super().handle_queue(url, kwdict)

    return JobCancelavel


# ---------------------------------------------------------------------------
# Normalizacao de entrada
# ---------------------------------------------------------------------------
def normalizar_perfil(texto: str) -> str:
    """Aceita '@fulano', 'fulano' ou a URL do perfil e devolve o username."""
    t = texto.strip().lstrip("@").strip()
    if not t:
        return ""
    m = re.search(r"instagram\.com/([^/?#]+)", t)
    if m:
        t = m.group(1)
    return t.strip("/").strip()


def url_da_secao(usuario: str, secao: str) -> str:
    base = f"https://www.instagram.com/{usuario}"
    if secao == "posts":
        return f"{base}/posts"
    if secao == "reels":
        return f"{base}/reels"
    if secao == "stories":
        return f"https://www.instagram.com/stories/{usuario}/"
    if secao == "highlights":
        return f"{base}/highlights"
    if secao == "tagged":
        return f"{base}/tagged"
    if secao == "avatar":
        return f"{base}/avatar"
    return base


def extrair_links(texto: str):
    """Extrai links de post / reel / tv / share de um bloco de texto."""
    achados = re.findall(
        r"https?://(?:www\.)?instagram\.com/[^\s\"'<>]+", texto, flags=re.I
    )
    limpos = []
    for u in achados:
        u = u.rstrip(".,;)")
        if u not in limpos:
            limpos.append(u)
    return limpos


# ---------------------------------------------------------------------------
# Configuracao do gallery-dl
# ---------------------------------------------------------------------------
def montar_config(destino, navegador, caminho_cookies, so_imagens, so_videos,
                  limite, incluir_fixados, salvar_json):
    gdl_config.clear()
    gdl_config.set((), "base-directory", str(destino))

    # Cookies (login)
    if navegador == ARQUIVO_COOKIES and caminho_cookies:
        gdl_config.set(("extractor",), "cookies", str(caminho_cookies))
    elif navegador not in (SEM_LOGIN, ARQUIVO_COOKIES):
        gdl_config.set(("extractor",), "cookies", [chave_navegador(navegador)])

    # Ritmo — evita bloqueio temporario do Instagram
    gdl_config.set(("extractor",), "sleep-request", [1.0, 2.5])
    gdl_config.set(("extractor",), "retries", 3)

    # Organizacao dos arquivos: pasta por usuario
    gdl_config.set(("extractor", "instagram"), "directory",
                   ["instagram", "{username}"])
    gdl_config.set(("extractor", "instagram"), "filename",
                   "{date:%Y-%m-%d}_{post_shortcode}_{num}.{extension}")

    # Filtros de tipo
    if so_imagens:
        gdl_config.set(("extractor", "instagram"), "videos", False)
    if so_videos:
        gdl_config.set(("extractor",), "image-filter",
                       "extension in ('mp4','mov','m4v')")

    if limite and limite > 0:
        gdl_config.set(("extractor", "instagram"), "max-posts", int(limite))

    gdl_config.set(("extractor", "instagram"), "pinned", bool(incluir_fixados))

    if salvar_json:
        gdl_config.set(("extractor",), "postprocessors",
                       [{"name": "metadata", "mode": "json"}])


# ---------------------------------------------------------------------------
# Planilha de metadados
# ---------------------------------------------------------------------------
CABECALHO = [
    "Usuario", "Data", "Tipo", "Link do post", "Legenda", "Curtidas",
    "Hashtags", "Mencoes", "Localizacao", "Arquivo", "Shortcode",
]


def coletar_metadados(url, log_fn):
    """Roda um DataJob (sem baixar midia) e devolve as linhas da planilha."""
    if CANCELAR.is_set():
        raise Cancelado()

    devnull = open(os.devnull, "w", encoding="utf-8")
    try:
        j = gdl_job.DataJob(url, file=devnull)
        j.run()
        dados = j.data
    finally:
        devnull.close()

    linhas = []
    vistos = set()
    for item in dados:
        if not isinstance(item, tuple) or len(item) < 3:
            continue
        tipo_msg = item[0]
        kw = item[-1]
        if not isinstance(kw, dict):
            continue
        # 3 = Message.Url (arquivo de midia)
        if tipo_msg != 3:
            continue

        short = kw.get("post_shortcode") or ""
        arquivo = f"{kw.get('filename', '')}.{kw.get('extension', '')}"
        chave = (short, arquivo)
        if chave in vistos:
            continue
        vistos.add(chave)

        data = kw.get("date")
        if isinstance(data, datetime):
            data_txt = data.strftime("%d/%m/%Y %H:%M")
        else:
            data_txt = str(data or "")

        legenda = (kw.get("description") or "").strip()
        tags = kw.get("tags") or []
        mencoes = kw.get("mentions") or []
        local = kw.get("location_slug") or ""

        linhas.append([
            kw.get("username") or (kw.get("owner_id") and str(kw["owner_id"])) or "",
            data_txt,
            kw.get("typename") or kw.get("type") or "",
            kw.get("post_url") or (f"https://www.instagram.com/p/{short}/" if short else ""),
            legenda,
            kw.get("likes", ""),
            ", ".join(tags) if isinstance(tags, list) else str(tags),
            ", ".join(mencoes) if isinstance(mencoes, list) else str(mencoes),
            local,
            arquivo,
            short,
        ])

    log_fn(f"  {len(linhas)} itens lidos de {url}")
    return linhas


def salvar_planilha(linhas, caminho):
    wb = Workbook()
    ws = wb.active
    ws.title = "Instagram"

    fill = PatternFill("solid", fgColor="1F2937")
    for i, titulo in enumerate(CABECALHO, start=1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(vertical="center")

    for linha in linhas:
        ws.append(linha)

    larguras = [18, 18, 12, 45, 70, 10, 30, 25, 20, 35, 15]
    for i, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CABECALHO))}{ws.max_row}"
    wb.save(caminho)


# ---------------------------------------------------------------------------
# Interface grafica
# ---------------------------------------------------------------------------
BG = "#f2f2f2"
FG = "#1a1a1a"


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME} v{VERSION}")
        self.geometry("880x720")
        self.minsize(820, 660)

        self._forcar_tema()

        self.fila = queue.Queue()
        self.thread = None
        self.caminho_cookies = tk.StringVar(value="")

        self._montar_ui()
        self.after(120, self._drenar_fila)
        self.after(300, self._redesenhar)

        if gdl_job is None:
            messagebox.showerror(
                APP_NAME,
                "A biblioteca gallery-dl nao foi encontrada.\n\n"
                "Instale com:  pip install gallery-dl",
            )

    def _forcar_tema(self):
        """O Tk 8.5 da Apple desenha janela preta no macOS moderno.
        Forcamos o tema 'clam' e cores explicitas para nao depender do sistema."""
        self.configure(bg=BG)
        estilo = ttk.Style(self)
        try:
            estilo.theme_use("clam")
        except tk.TclError:
            pass
        estilo.configure(".", background=BG, foreground=FG, fieldbackground="#ffffff")
        estilo.configure("TFrame", background=BG)
        estilo.configure("TLabel", background=BG, foreground=FG)
        estilo.configure("TLabelframe", background=BG, foreground=FG)
        estilo.configure("TLabelframe.Label", background=BG, foreground=FG)
        estilo.configure("TCheckbutton", background=BG, foreground=FG)
        estilo.configure("TButton", background="#e4e4e4", foreground=FG, padding=5)
        estilo.map("TButton", background=[("active", "#d0d0d0")])
        estilo.configure("TNotebook", background=BG)
        estilo.configure("TNotebook.Tab", background="#dcdcdc", foreground=FG, padding=(12, 6))
        estilo.map("TNotebook.Tab", background=[("selected", BG)])
        estilo.configure("TEntry", fieldbackground="#ffffff", foreground=FG)
        estilo.configure("TCombobox", fieldbackground="#ffffff", foreground=FG)
        estilo.configure("TSpinbox", fieldbackground="#ffffff", foreground=FG)
        estilo.configure("TProgressbar", background="#2563eb", troughcolor="#dcdcdc")

    def _redesenhar(self):
        """Chuta a janela para o Tk 8.5 pintar o conteudo (bug da janela preta)."""
        try:
            g = self.geometry()
            larg, resto = g.split("x", 1)
            alt = resto.split("+", 1)[0]
            self.geometry(f"{int(larg)}x{int(alt) + 1}")
            self.update_idletasks()
            self.geometry(g)
            self.update_idletasks()
        except Exception:  # noqa: BLE001
            pass

    # -- construcao da interface -------------------------------------------
    def _montar_ui(self):
        pad = {"padx": 10, "pady": 6}

        topo = ttk.Frame(self)
        topo.pack(fill="x", **pad)
        ttk.Label(topo, text=APP_NAME, font=("Segoe UI", 15, "bold")).pack(side="left")
        ttk.Label(
            topo,
            text="baixa posts, carrosseis, reels, stories e metadados",
            foreground="#666",
        ).pack(side="left", padx=10)

        # --- Destino ---
        box_dest = ttk.LabelFrame(self, text="Pasta de destino")
        box_dest.pack(fill="x", **pad)
        self.destino = tk.StringVar(value=str(Path.home() / "Downloads" / "Instagram"))
        ttk.Entry(box_dest, textvariable=self.destino).pack(
            side="left", fill="x", expand=True, padx=8, pady=8
        )
        ttk.Button(box_dest, text="Escolher...", command=self._escolher_destino).pack(
            side="left", padx=8, pady=8
        )
        ttk.Button(box_dest, text="Abrir pasta", command=self._abrir_destino).pack(
            side="left", padx=(0, 8), pady=8
        )

        # --- Login ---
        box_login = ttk.LabelFrame(self, text="Login (o Instagram exige conta logada)")
        box_login.pack(fill="x", **pad)
        ttk.Label(box_login, text="Puxar cookies de:").pack(side="left", padx=8, pady=8)
        self.navegador = tk.StringVar(value="Chrome")
        cb = ttk.Combobox(
            box_login, textvariable=self.navegador, values=NAVEGADORES,
            state="readonly", width=26,
        )
        cb.pack(side="left", pady=8)
        cb.bind("<<ComboboxSelected>>", self._trocou_navegador)
        self.btn_cookies = ttk.Button(
            box_login, text="Selecionar cookies.txt", command=self._escolher_cookies,
            state="disabled",
        )
        self.btn_cookies.pack(side="left", padx=8, pady=8)
        ttk.Button(box_login, text="Testar login", command=self._testar_login).pack(
            side="left", padx=4, pady=8
        )
        self.lbl_cookies = ttk.Label(box_login, text="", foreground="#666")
        self.lbl_cookies.pack(side="left", padx=4)

        # --- Abas ---
        self.abas = ttk.Notebook(self)
        self.abas.pack(fill="both", expand=True, **pad)
        self._aba_links()
        self._aba_perfil()

        # --- Opcoes gerais ---
        box_op = ttk.LabelFrame(self, text="Opcoes")
        box_op.pack(fill="x", **pad)
        self.so_imagens = tk.BooleanVar(value=False)
        self.so_videos = tk.BooleanVar(value=False)
        self.salvar_json = tk.BooleanVar(value=False)
        self.gerar_planilha = tk.BooleanVar(value=True)
        self.baixar_midia = tk.BooleanVar(value=True)

        ttk.Checkbutton(box_op, text="Baixar midia", variable=self.baixar_midia).grid(
            row=0, column=0, sticky="w", padx=8, pady=4
        )
        ttk.Checkbutton(
            box_op, text="Gerar planilha de metadados (.xlsx)",
            variable=self.gerar_planilha,
        ).grid(row=0, column=1, sticky="w", padx=8, pady=4)
        ttk.Checkbutton(box_op, text="Salvar .json de cada post", variable=self.salvar_json).grid(
            row=0, column=2, sticky="w", padx=8, pady=4
        )
        ttk.Checkbutton(box_op, text="So imagens", variable=self.so_imagens).grid(
            row=1, column=0, sticky="w", padx=8, pady=4
        )
        ttk.Checkbutton(box_op, text="So videos", variable=self.so_videos).grid(
            row=1, column=1, sticky="w", padx=8, pady=4
        )

        # --- Acoes ---
        box_acao = ttk.Frame(self)
        box_acao.pack(fill="x", **pad)
        self.btn_rodar = ttk.Button(box_acao, text="Coletar", command=self._rodar)
        self.btn_rodar.pack(side="left")
        self.btn_cancelar = ttk.Button(
            box_acao, text="Cancelar", command=self._cancelar, state="disabled"
        )
        self.btn_cancelar.pack(side="left", padx=8)

        self.barra = ttk.Progressbar(box_acao, mode="determinate", length=280)
        self.barra.pack(side="left", padx=16, fill="x", expand=True)
        self.status = ttk.Label(box_acao, text="Pronto.", foreground="#444")
        self.status.pack(side="left", padx=8)

        # --- Log ---
        box_log = ttk.LabelFrame(self, text="Registro")
        box_log.pack(fill="both", expand=True, **pad)
        self.log = tk.Text(box_log, height=10, wrap="word", state="disabled",
                           font=("Menlo", 10), bg="#ffffff", fg=FG,
                           insertbackground=FG, highlightthickness=1,
                           highlightbackground="#c0c0c0")
        sb = ttk.Scrollbar(box_log, command=self.log.yview)
        self.log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.log.pack(fill="both", expand=True, padx=6, pady=6)

    def _aba_links(self):
        f = ttk.Frame(self.abas)
        self.abas.add(f, text="  Links (post / carrossel / reel)  ")
        ttk.Label(
            f,
            text="Cole um link por linha. Aceita /p/, /reel/, /tv/ e link de compartilhamento.\n"
                 "Carrossel baixa todas as fotos e videos automaticamente.",
            foreground="#555",
        ).pack(anchor="w", padx=10, pady=(10, 4))
        self.txt_links = tk.Text(f, height=8, wrap="none", font=("Menlo", 10),
                                 bg="#ffffff", fg=FG, insertbackground=FG,
                                 highlightthickness=1, highlightbackground="#c0c0c0")
        self.txt_links.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _aba_perfil(self):
        f = ttk.Frame(self.abas)
        self.abas.add(f, text="  Perfil inteiro  ")

        linha = ttk.Frame(f)
        linha.pack(fill="x", padx=10, pady=(12, 6))
        ttk.Label(linha, text="Perfil:").pack(side="left")
        self.perfil = tk.StringVar()
        ttk.Entry(linha, textvariable=self.perfil, width=32).pack(side="left", padx=8)
        ttk.Label(linha, text="(ex.: @mvitorino_ ou o link do perfil)",
                  foreground="#666").pack(side="left")

        ttk.Label(f, text="O que baixar:", font=("Segoe UI", 9, "bold")).pack(
            anchor="w", padx=10, pady=(10, 2)
        )
        grade = ttk.Frame(f)
        grade.pack(anchor="w", padx=10)
        self.secoes = {}
        for i, (chave, rotulo) in enumerate(SECOES_PERFIL):
            v = tk.BooleanVar(value=(chave == "posts"))
            self.secoes[chave] = v
            ttk.Checkbutton(grade, text=rotulo, variable=v).grid(
                row=i // 2, column=i % 2, sticky="w", padx=6, pady=3
            )

        linha2 = ttk.Frame(f)
        linha2.pack(fill="x", padx=10, pady=(12, 6))
        ttk.Label(linha2, text="Limite de posts (0 = tudo):").pack(side="left")
        self.limite = tk.IntVar(value=30)
        ttk.Spinbox(linha2, from_=0, to=5000, textvariable=self.limite,
                    width=8).pack(side="left", padx=8)
        self.fixados = tk.BooleanVar(value=True)
        ttk.Checkbutton(linha2, text="Incluir posts fixados",
                        variable=self.fixados).pack(side="left", padx=16)

        ttk.Label(
            f,
            text="Aviso: perfis privados exigem que a conta logada siga o perfil.\n"
                 "Baixar muita coisa de uma vez pode gerar bloqueio temporario do Instagram.",
            foreground="#8a5a00",
        ).pack(anchor="w", padx=10, pady=(14, 8))

    # -- eventos de UI -----------------------------------------------------
    def _trocou_navegador(self, _=None):
        arquivo = self.navegador.get() == ARQUIVO_COOKIES
        self.btn_cookies.configure(state="normal" if arquivo else "disabled")
        if not arquivo:
            self.lbl_cookies.configure(text="")

    def _testar_login(self):
        if gdl_cookies is None:
            messagebox.showerror(APP_NAME, "gallery-dl nao instalado.")
            return
        self.status.configure(text="Testando login...")
        self.update_idletasks()
        try:
            ok, msg = testar_login(self.navegador.get(), self.caminho_cookies.get())
        except Exception as e:  # noqa: BLE001
            ok, msg = False, f"Erro no teste: {e}"
        self.status.configure(text="Login OK." if ok else "Login falhou.")
        self._escrever(("[login OK] " if ok else "[login FALHOU] ") + msg.replace("\n", " "))
        if ok:
            messagebox.showinfo(APP_NAME, msg)
        else:
            messagebox.showwarning(APP_NAME, msg)

    def _escolher_cookies(self):
        p = filedialog.askopenfilename(
            title="Selecione o cookies.txt",
            filetypes=[("Texto", "*.txt"), ("Todos", "*.*")],
        )
        if p:
            self.caminho_cookies.set(p)
            self.lbl_cookies.configure(text=Path(p).name)

    def _escolher_destino(self):
        p = filedialog.askdirectory(title="Pasta de destino")
        if p:
            self.destino.set(p)

    def _abrir_destino(self):
        d = Path(self.destino.get())
        d.mkdir(parents=True, exist_ok=True)
        if sys.platform.startswith("win"):
            os.startfile(str(d))  # noqa: S606
        elif sys.platform == "darwin":
            os.system(f'open "{d}"')
        else:
            webbrowser.open(str(d))

    def _escrever(self, txt):
        self.fila.put(("log", txt))

    def _drenar_fila(self):
        try:
            while True:
                tipo, valor = self.fila.get_nowait()
                if tipo == "log":
                    self.log.configure(state="normal")
                    self.log.insert("end", valor + "\n")
                    self.log.see("end")
                    self.log.configure(state="disabled")
                elif tipo == "status":
                    self.status.configure(text=valor)
                elif tipo == "barra":
                    self.barra["value"] = valor
                elif tipo == "max":
                    self.barra["maximum"] = max(1, valor)
                elif tipo == "fim":
                    self._terminou(valor)
        except queue.Empty:
            pass
        self.after(120, self._drenar_fila)

    def _cancelar(self):
        CANCELAR.set()
        self.fila.put(("status", "Cancelando..."))
        self._escrever("Cancelamento pedido. Aguardando o item atual terminar...")

    def _terminou(self, msg):
        self.btn_rodar.configure(state="normal")
        self.btn_cancelar.configure(state="disabled")
        self.status.configure(text=msg)
        self.barra["value"] = self.barra["maximum"]

    # -- execucao ----------------------------------------------------------
    def _rodar(self):
        if gdl_job is None:
            messagebox.showerror(APP_NAME, "gallery-dl nao instalado.")
            return
        if self.gerar_planilha.get() and Workbook is None:
            messagebox.showerror(APP_NAME, "openpyxl nao instalado (necessario para a planilha).")
            return
        if not self.baixar_midia.get() and not self.gerar_planilha.get():
            messagebox.showwarning(APP_NAME, "Marque pelo menos 'Baixar midia' ou 'Gerar planilha'.")
            return

        aba = self.abas.index(self.abas.select())
        if aba == 0:
            urls = extrair_links(self.txt_links.get("1.0", "end"))
            if not urls:
                messagebox.showwarning(APP_NAME, "Cole pelo menos um link do Instagram.")
                return
        else:
            usuario = normalizar_perfil(self.perfil.get())
            if not usuario:
                messagebox.showwarning(APP_NAME, "Digite o perfil.")
                return
            escolhidas = [k for k, v in self.secoes.items() if v.get()]
            if not escolhidas:
                messagebox.showwarning(APP_NAME, "Marque pelo menos uma secao do perfil.")
                return
            urls = [url_da_secao(usuario, s) for s in escolhidas]

        CANCELAR.clear()
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")
        self.btn_rodar.configure(state="disabled")
        self.btn_cancelar.configure(state="normal")
        self.barra["value"] = 0
        self.fila.put(("max", len(urls) * 2))

        self.thread = threading.Thread(target=self._trabalho, args=(urls,), daemon=True)
        self.thread.start()

    def _trabalho(self, urls):
        ponte = instalar_ponte_de_log(self._escrever)
        try:
            destino = Path(self.destino.get())
            destino.mkdir(parents=True, exist_ok=True)

            # Confere o login antes de gastar tempo.
            ok, msg = testar_login(self.navegador.get(), self.caminho_cookies.get())
            if ok:
                self._escrever(f"[login] {msg}")
            else:
                self._escrever(f"[login] ATENCAO: {msg}")

            montar_config(
                destino=destino,
                navegador=self.navegador.get(),
                caminho_cookies=self.caminho_cookies.get(),
                so_imagens=self.so_imagens.get(),
                so_videos=self.so_videos.get(),
                limite=self.limite.get(),
                incluir_fixados=self.fixados.get(),
                salvar_json=self.salvar_json.get(),
            )

            self._escrever(f"Destino: {destino}")
            self._escrever(f"Login: cookies de {self.navegador.get()}")
            self._escrever("-" * 60)

            baixados = [0]

            def progresso(nome):
                baixados[0] += 1
                self.fila.put(("status", f"{baixados[0]} arquivo(s)"))
                self._escrever(f"  baixado: {nome}")

            passo = 0
            linhas_planilha = []

            for url in urls:
                if CANCELAR.is_set():
                    raise Cancelado()

                # --- metadados ---
                if self.gerar_planilha.get():
                    self.fila.put(("status", "Lendo metadados..."))
                    self._escrever(f"[metadados] {url}")
                    try:
                        linhas_planilha += coletar_metadados(url, self._escrever)
                    except Cancelado:
                        raise
                    except Exception as e:  # noqa: BLE001
                        self._escrever(f"  ERRO nos metadados: {e}")
                passo += 1
                self.fila.put(("barra", passo))

                # --- midia ---
                if self.baixar_midia.get():
                    self.fila.put(("status", "Baixando..."))
                    self._escrever(f"[download] {url}")
                    try:
                        JobCls = _fabricar_job_download(self._escrever, progresso)
                        JobCls(url).run()
                    except Cancelado:
                        raise
                    except gdl_exception.GalleryDLException as e:
                        self._escrever(f"  ERRO: {e}")
                    except Exception as e:  # noqa: BLE001
                        self._escrever(f"  ERRO: {e}")
                passo += 1
                self.fila.put(("barra", passo))

            # --- salva planilha ---
            if self.gerar_planilha.get() and linhas_planilha:
                carimbo = datetime.now().strftime("%Y-%m-%d_%H%M")
                arq = destino / f"instagram_metadados_{carimbo}.xlsx"
                salvar_planilha(linhas_planilha, arq)
                self._escrever("-" * 60)
                self._escrever(f"Planilha salva: {arq}")
            elif self.gerar_planilha.get():
                self._escrever("Nenhum metadado coletado (verifique o login).")

            self._escrever("-" * 60)
            self._escrever(f"Concluido. {baixados[0]} arquivo(s) baixado(s).")
            self.fila.put(("fim", f"Concluido — {baixados[0]} arquivo(s)"))

        except Cancelado:
            self._escrever("Cancelado pelo usuario.")
            self.fila.put(("fim", "Cancelado."))
        except Exception:  # noqa: BLE001
            self._escrever("ERRO INESPERADO:\n" + traceback.format_exc())
            self.fila.put(("fim", "Erro — veja o registro."))
        finally:
            logging.getLogger().removeHandler(ponte)


def main():
    App().mainloop()


if __name__ == "__main__":
    main()
