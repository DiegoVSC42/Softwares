#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Normalizador de Números — formata telefones brasileiros em lote.

GUI em Tkinter. Cole números (um por linha) ou carregue um arquivo
TXT/CSV/XLSX, escolha o formato de saída e normalize:
- remove tudo que não é dígito;
- trata o código do país (55);
- aplica DDD padrão quando o número vem sem DDD (opcional);
- acrescenta o 9º dígito em celulares antigos (8 dígitos começando em 6-9);
- remove duplicados (opcional);
- separa os inválidos com o motivo.
"""

import csv
import os
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

FORMATOS = {
    "WhatsApp (5562999998888)": "whatsapp",
    "Internacional (+55 62 99999-8888)": "internacional",
    "Nacional ((62) 99999-8888)": "nacional",
    "Somente dígitos com DDD (62999998888)": "digitos",
}


def normalizar_numero(bruto, ddd_padrao=""):
    """Retorna (ddd, numero) ou (None, motivo)."""
    d = re.sub(r"\D", "", str(bruto))
    if not d:
        return None, "vazio"
    d = d.lstrip("0")
    # remove código do país, se presente
    if d.startswith("55") and len(d) in (12, 13):
        d = d[2:]
    if len(d) in (8, 9):
        if not ddd_padrao:
            return None, "sem DDD (informe um DDD padrão)"
        d = ddd_padrao + d
    if len(d) == 10:
        ddd, resto = d[:2], d[2:]
        if resto[0] in "6789":  # celular antigo sem o 9
            resto = "9" + resto
    elif len(d) == 11:
        ddd, resto = d[:2], d[2:]
        if resto[0] != "9" and resto[0] not in "2345":
            return None, "número de 11 dígitos inválido"
    else:
        return None, f"quantidade de dígitos inválida ({len(d)})"
    if not (11 <= int(ddd) <= 99) or ddd[1] == "0":
        return None, f"DDD inválido ({ddd})"
    return ddd, resto


def formatar(ddd, resto, formato):
    meio = len(resto) - 4
    if formato == "whatsapp":
        return f"55{ddd}{resto}"
    if formato == "internacional":
        return f"+55 {ddd} {resto[:meio]}-{resto[meio:]}"
    if formato == "nacional":
        return f"({ddd}) {resto[:meio]}-{resto[meio:]}"
    return f"{ddd}{resto}"


def ler_tabela(caminho):
    """Lê CSV ou XLSX e devolve a lista de linhas (cada linha = lista de células)."""
    ext = os.path.splitext(caminho)[1].lower()
    linhas = []
    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            linhas.append(["" if c is None else str(c) for c in row])
        wb.close()
    else:
        with open(caminho, newline="", encoding="utf-8", errors="replace") as f:
            amostra = f.read(4096)
            f.seek(0)
            try:
                dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t")
            except csv.Error:
                dialeto = csv.excel
            for row in csv.reader(f, dialeto):
                linhas.append(row)
    return linhas


def ler_txt(caminho):
    with open(caminho, encoding="utf-8", errors="replace") as f:
        return [l.strip() for l in f if l.strip()]


class EscolhaColuna(tk.Toplevel):
    """Janela para escolher qual coluna da planilha tem os números."""

    def __init__(self, parent, tabela):
        super().__init__(parent)
        self.title("Qual coluna tem os números?")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.escolha = None
        self.tem_cabecalho = tk.BooleanVar(value=True)

        ncols = max(len(l) for l in tabela)
        cabecalho = tabela[0] + [""] * (ncols - len(tabela[0]))
        self.opcoes = []
        for i in range(ncols):
            nome = cabecalho[i].strip() or f"Coluna {i + 1}"
            amostras = [l[i] for l in tabela[1:6] if i < len(l) and str(l[i]).strip()]
            exemplo = " · ".join(str(a)[:18] for a in amostras[:3])
            self.opcoes.append(f"{nome}  ({exemplo})" if exemplo else nome)

        ttk.Label(self, text="Escolha a coluna com os números de telefone:",
                  padding=(12, 12, 12, 4)).pack(anchor="w")
        self.lista = tk.Listbox(self, height=min(12, ncols), width=60, exportselection=False)
        for op in self.opcoes:
            self.lista.insert("end", op)
        # pré-seleciona a coluna que mais parece telefone
        melhor, melhor_qtd = 0, -1
        for i in range(ncols):
            qtd = sum(1 for l in tabela[1:30]
                      if i < len(l) and re.search(r"\d{8}", re.sub(r"\D", "", str(l[i]))))
            if qtd > melhor_qtd:
                melhor, melhor_qtd = i, qtd
        self.lista.selection_set(melhor)
        self.lista.see(melhor)
        self.lista.pack(padx=12, fill="x")
        self.lista.bind("<Double-Button-1>", lambda e: self.confirmar())

        ttk.Checkbutton(self, text="A primeira linha é cabeçalho (ignorar)",
                        variable=self.tem_cabecalho).pack(anchor="w", padx=12, pady=4)
        rodape = ttk.Frame(self, padding=12)
        rodape.pack(fill="x")
        ttk.Button(rodape, text="Usar esta coluna", command=self.confirmar).pack(side="right")
        ttk.Button(rodape, text="Cancelar", command=self.destroy).pack(side="right", padx=6)

    def confirmar(self):
        sel = self.lista.curselection()
        if sel:
            self.escolha = sel[0]
        self.destroy()


class App:
    def __init__(self, root):
        self.root = root
        root.title("Normalizador de Números")
        root.geometry("820x640")
        self.cancelar = threading.Event()

        topo = ttk.Frame(root, padding=10)
        topo.pack(fill="x")

        ttk.Button(topo, text="Carregar arquivo (TXT/CSV/XLSX)", command=self.carregar).pack(side="left")
        ttk.Label(topo, text="   DDD padrão:").pack(side="left")
        self.ddd_var = tk.StringVar()
        ttk.Entry(topo, textvariable=self.ddd_var, width=4).pack(side="left")
        ttk.Label(topo, text="   Formato:").pack(side="left")
        self.fmt_var = tk.StringVar(value=list(FORMATOS)[0])
        ttk.Combobox(topo, textvariable=self.fmt_var, values=list(FORMATOS), state="readonly", width=34).pack(side="left")
        self.dup_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(topo, text="Remover duplicados", variable=self.dup_var).pack(side="left", padx=8)
        self.manter_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            topo,
            text="Manter inválidos (saída alinhada com a entrada)",
            variable=self.manter_var,
        ).pack(side="left")

        corpo = ttk.PanedWindow(root, orient="horizontal")
        corpo.pack(fill="both", expand=True, padx=10)

        esq = ttk.Frame(corpo)
        ttk.Label(esq, text="Números de entrada (um por linha):").pack(anchor="w")
        self.entrada = tk.Text(esq, width=40)
        self.entrada.pack(fill="both", expand=True)
        corpo.add(esq, weight=1)

        dire = ttk.Frame(corpo)
        ttk.Label(dire, text="Resultado:").pack(anchor="w")
        self.saida = tk.Text(dire, width=40)
        self.saida.pack(fill="both", expand=True)
        corpo.add(dire, weight=1)

        meio = ttk.Frame(root, padding=10)
        meio.pack(fill="x")
        self.btn_ok = ttk.Button(meio, text="Normalizar", command=self.iniciar)
        self.btn_ok.pack(side="left")
        self.btn_cancel = ttk.Button(meio, text="Cancelar", command=self.cancelar.set, state="disabled")
        self.btn_cancel.pack(side="left", padx=6)
        ttk.Button(meio, text="Copiar resultado", command=self.copiar).pack(side="left", padx=6)
        ttk.Button(meio, text="Salvar resultado...", command=self.salvar).pack(side="left")

        self.progresso = ttk.Progressbar(root, mode="determinate")
        self.progresso.pack(fill="x", padx=10)
        self.status = tk.StringVar(value="Pronto.")
        ttk.Label(root, textvariable=self.status, padding=6).pack(anchor="w")

        ttk.Label(root, text="Inválidos:").pack(anchor="w", padx=10)
        self.invalidos = tk.Text(root, height=6, foreground="#a33")
        self.invalidos.pack(fill="x", padx=10, pady=(0, 10))

    def carregar(self):
        caminho = filedialog.askopenfilename(
            filetypes=[("Listas", "*.txt *.csv *.xlsx"), ("Todos", "*.*")])
        if not caminho:
            return
        ext = os.path.splitext(caminho)[1].lower()
        try:
            if ext in (".csv", ".xlsx"):
                tabela = ler_tabela(caminho)
                # descarta apenas as linhas totalmente vazias do FINAL (mantém as do meio,
                # para a saída alinhada não desalinhar com a planilha)
                while tabela and not any(str(c).strip() for c in tabela[-1]):
                    tabela.pop()
                if not tabela:
                    messagebox.showwarning("Arquivo vazio", "Não encontrei dados no arquivo.")
                    return
                dlg = EscolhaColuna(self.root, tabela)
                self.root.wait_window(dlg)
                if dlg.escolha is None:
                    self.status.set("Carregamento cancelado.")
                    return
                col = dlg.escolha
                dados = tabela[1:] if dlg.tem_cabecalho.get() else tabela
                linhas = [str(l[col]).strip() if col < len(l) else "" for l in dados]
            else:
                linhas = ler_txt(caminho)
        except Exception as e:
            messagebox.showerror("Erro ao ler o arquivo", str(e))
            return
        self.entrada.delete("1.0", "end")
        self.entrada.insert("1.0", "\n".join(linhas))
        self.status.set(f"{len(linhas)} linhas carregadas de {os.path.basename(caminho)}.")

    def iniciar(self):
        ddd = self.ddd_var.get().strip()
        if ddd and not re.fullmatch(r"[1-9][1-9]", ddd):
            messagebox.showwarning("DDD inválido", "O DDD padrão deve ter 2 dígitos (ex.: 62).")
            return
        self.cancelar.clear()
        self.btn_ok.config(state="disabled")
        self.btn_cancel.config(state="normal")
        threading.Thread(target=self.processar, daemon=True).start()

    def processar(self):
        alinhado = self.manter_var.get()
        todas = self.entrada.get("1.0", "end").splitlines()
        # no modo alinhado, cada linha da entrada (mesmo vazia) gera uma linha na saída
        linhas = todas if alinhado else [l for l in todas if l.strip()]
        while alinhado and linhas and not linhas[-1].strip():
            linhas.pop()  # descarta só as vazias do final
        fmt = FORMATOS[self.fmt_var.get()]
        ddd_padrao = self.ddd_var.get().strip()
        total = len(linhas) or 1
        vistos, saida, invalidos = set(), [], []
        for i, linha in enumerate(linhas):
            if self.cancelar.is_set():
                break
            if alinhado and not linha.strip():
                saida.append("")
                continue
            r = normalizar_numero(linha, ddd_padrao)
            if r[0] is None:
                invalidos.append(f"linha {i + 1}: {linha}  →  {r[1]}")
                if alinhado:
                    saida.append(linha)  # inválido permanece como estava, na mesma linha
            else:
                chave = r[0] + r[1]
                if not alinhado and self.dup_var.get() and chave in vistos:
                    pass  # duplicado removido (só fora do modo alinhado)
                else:
                    vistos.add(chave)
                    saida.append(formatar(r[0], r[1], fmt))
            if i % 200 == 0:
                self.root.after(0, self.progresso.config, {"value": 100 * i / total})
        self.root.after(0, self.mostrar, saida, invalidos, len(linhas), alinhado)

    def mostrar(self, saida, invalidos, total, alinhado=False):
        self.saida.delete("1.0", "end")
        self.saida.insert("1.0", "\n".join(saida))
        self.invalidos.delete("1.0", "end")
        self.invalidos.insert("1.0", "\n".join(invalidos))
        self.progresso.config(value=100)
        self.btn_ok.config(state="normal")
        self.btn_cancel.config(state="disabled")
        corrigidos = len([s for s in saida if s.strip()]) - len(invalidos) if alinhado else len(saida)
        extra = " · saída alinhada linha a linha, pronta para colar na planilha" if alinhado else ""
        self.status.set(f"{total} lidos · {corrigidos} corrigidos · {len(invalidos)} inválidos mantidos{extra}."
                        if alinhado else
                        f"{total} lidos · {corrigidos} válidos · {len(invalidos)} inválidos.")

    def copiar(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.saida.get("1.0", "end").strip())
        self.status.set("Resultado copiado para a área de transferência.")

    def salvar(self):
        caminho = filedialog.asksaveasfilename(
            defaultextension=".txt", initialfile="numeros-normalizados.txt",
            filetypes=[("Texto", "*.txt"), ("CSV", "*.csv")])
        if not caminho:
            return
        with open(caminho, "w", encoding="utf-8") as f:
            f.write(self.saida.get("1.0", "end").strip() + "\n")
        self.status.set(f"Salvo em {caminho}.")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
